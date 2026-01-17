from __future__ import annotations

from openpyxl import load_workbook
from datetime import datetime, date, timedelta
from pathlib import Path
import csv

# =====================
# CONFIG (valeurs par defaut)
# =====================
BASE_DIR = Path(__file__).resolve().parent

# Fichier Excel par defaut (relatif au script)
DEFAULT_XLSX_PATH = BASE_DIR / "ROW 2026 FR113 NextGen Schedule 4.1_.xlsx"
SHEET_NAME = None

# Lignes fixes du planning
ROW_MONTH = 1
ROW_DAY = 4          # jour du mois / date
ROW_WEEKDAY = 3      # Mon/Tue/Wed/Sat/Sun...

# Ligne cible determinee via la colonne 1 (ex: 'M09')
DEFAULT_TARGET_LABEL = "M09"

# Codes
ASTREINTE_CODES = {"-14", "-X"}
BRIDGE_CODE = "H"
CONGES_CODES = {"F", "V"}

SAMEDI_TRAVAIL_CODE = 8.5
SAMEDI_TRAVAIL_WEEKDAY = "sat"  # detection 'Sat' / 'Saturday'

# Exports
EXPORT_CSV = True
CSV_OUT_BASE = "exportcsv"

EXPORT_ICS = True
ICS_OUT_BASE = "exportcalendrier"
ICS_CALNAME = "Astreintes"
ICS_TIMEZONE = "Europe/Paris"

# Calendrier: 2 options basiques
# - per_category: 1 fichier ICS par categorie
# - all_in_one: 1 fichier ICS "tout"
DEFAULT_ICS_MODE = "per_category"  # "per_category" ou "all_in_one"

ICS_EVENT_TITLE_ASTREINTE = "Astreinte"
ICS_EVENT_TITLE_CONGES = "Conges"
ICS_EVENT_TITLE_SAMEDI_TRAV = "Samedi travaille"
ICS_EVENT_TITLE_REPOS_HEBDO = "Repos hebdomadaire"

INCLUDE_SINGLE_DAY_CONGES = True

# =====================
# HELPERS
# =====================
FR_DAYS = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]


def fmt(d: date) -> str:
    return f"{FR_DAYS[d.weekday()]} {d.strftime('%d/%m/%y')}"


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def norm(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return str(v).replace(".0", "")
    return str(v).strip()


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "."))
    except Exception:
        return None


def is_weekday_sat(v) -> bool:
    return v is not None and str(v).strip().lower().startswith(SAMEDI_TRAVAIL_WEEKDAY)


def is_samedi_trav_code(v) -> bool:
    f = to_float(v)
    return f is not None and abs(f - SAMEDI_TRAVAIL_CODE) < 1e-9


def is_weekend_label(v) -> bool:
    if v is None:
        return False
    s = str(v).strip().lower()
    return s.startswith("sat") or s.startswith("sun")


def is_repos_hebdo_code(v) -> bool:
    if v is None:
        return False
    return str(v).strip().upper() == "X"


def group_consecutive_days(days: set[date]):
    if not days:
        return []
    ds = sorted(days)
    blocks = []
    start = prev = ds[0]
    for d in ds[1:]:
        if (d - prev).days == 1:
            prev = d
        else:
            blocks.append((start, prev))
            start = prev = d
    blocks.append((start, prev))
    return blocks


def filter_conges_blocks(blocks, include_single):
    return blocks if include_single else [b for b in blocks if b[0] != b[1]]


def ics_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,").replace("\n", "\\n")


def block_uid(prefix: str, start: date, end: date) -> str:
    return f"{prefix}-{start.isoformat()}_{end.isoformat()}@local"


def find_row_target(ws, target_label: str) -> int:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() == target_label:
            return r
    raise ValueError(f"Ligne cible introuvable : valeur '{target_label}' non trouvee en colonne 1")


# =====================
# CORE
# =====================

def compute_from_excel(xlsx_path: Path, target_label: str):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

    row_target = find_row_target(ws, target_label)

    max_col = ws.max_column
    date_code: dict[date, str] = {}
    unknown_cols: list[int] = []

    samedi_trav_days: set[date] = set()
    repos_hebdo_days: set[date] = set()

    for col in range(1, max_col + 1):
        code_raw = ws.cell(row=row_target, column=col).value
        code = norm(code_raw)
        if not code:
            continue

        d = to_date(ws.cell(row=ROW_DAY, column=col).value)
        if d is None:
            day = ws.cell(row=ROW_DAY, column=col).value
            month = ws.cell(row=ROW_MONTH, column=col).value
            if isinstance(day, (int, float)) and isinstance(month, (int, float)):
                try:
                    d = date(2026, int(month), int(day))
                except Exception:
                    pass

        if d is None:
            unknown_cols.append(col)
            continue

        wd = ws.cell(row=ROW_WEEKDAY, column=col).value

        # Samedi travaille : 8,5 sur un 'Sat' -> evenement le samedi (meme date)
        if is_weekday_sat(wd) and is_samedi_trav_code(code_raw):
            samedi_trav_days.add(d)

        # Repos hebdomadaire : si 'X' et que le libelle jour n'est pas Sat/Sun
        if is_repos_hebdo_code(code_raw) and not is_weekend_label(wd):
            repos_hebdo_days.add(d)

        # date_code standard (astreintes/conges/H)
        priority = {"-14": 3, "-X": 3, "H": 2, "F": 1, "V": 1}
        prev = date_code.get(d)
        if prev is None or priority.get(code, 0) > priority.get(prev, 0):
            date_code[d] = code

    # Sets
    sorted_days = sorted(date_code.items())
    astreinte_days: set[date] = set()
    conges_days: set[date] = set()

    for d, c in sorted_days:
        if c in ASTREINTE_CODES:
            astreinte_days.add(d)
        if c in CONGES_CODES:
            conges_days.add(d)

    # H pont
    seg = []
    segments = []
    prev_d = None
    for d, c in sorted_days:
        if prev_d is None or (d - prev_d).days == 1:
            seg.append((d, c))
        else:
            segments.append(seg)
            seg = [(d, c)]
        prev_d = d
    if seg:
        segments.append(seg)

    for seg in segments:
        i = 0
        while i < len(seg):
            if seg[i][1] in ASTREINTE_CODES:
                j = i + 1
                while j < len(seg) and seg[j][1] == BRIDGE_CODE:
                    j += 1
                if j < len(seg) and seg[j][1] in ASTREINTE_CODES:
                    for k in range(i + 1, j):
                        astreinte_days.add(seg[k][0])
                i = j
            else:
                i += 1

    # Blocks
    astreinte_blocks = group_consecutive_days(astreinte_days)
    conges_blocks = filter_conges_blocks(group_consecutive_days(conges_days), INCLUDE_SINGLE_DAY_CONGES)
    samedi_trav_blocks = group_consecutive_days(samedi_trav_days)
    repos_hebdo_blocks = group_consecutive_days(repos_hebdo_days)

    return {
        "xlsx_path": xlsx_path,
        "target_label": target_label,
        "unknown_cols": unknown_cols,
        "astreinte_days": astreinte_days,
        "conges_days": conges_days,
        "samedi_trav_days": samedi_trav_days,
        "repos_hebdo_days": repos_hebdo_days,
        "astreinte_blocks": astreinte_blocks,
        "conges_blocks": conges_blocks,
        "samedi_trav_blocks": samedi_trav_blocks,
        "repos_hebdo_blocks": repos_hebdo_blocks,
    }


def export_csv(result: dict, timestamp: str) -> Path:
    csv_out = Path(f"{CSV_OUT_BASE}_{timestamp}.csv")

    astreinte_days = result["astreinte_days"]
    conges_days = result["conges_days"]
    samedi_trav_days = result["samedi_trav_days"]
    repos_hebdo_days = result["repos_hebdo_days"]

    astreinte_blocks = result["astreinte_blocks"]
    conges_blocks = result["conges_blocks"]
    samedi_trav_blocks = result["samedi_trav_blocks"]
    repos_hebdo_blocks = result["repos_hebdo_blocks"]

    with csv_out.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["# SUMMARY"])
        w.writerow(["categorie", "jours", "blocs"])
        w.writerow(["astreinte", len(astreinte_days), len(astreinte_blocks)])
        w.writerow(["conges", len(conges_days), len(conges_blocks)])
        w.writerow(["samedi_travaille", len(samedi_trav_days), len(samedi_trav_blocks)])
        w.writerow(["repos_hebdo", len(repos_hebdo_days), len(repos_hebdo_blocks)])
        w.writerow([])
        w.writerow(["# DAYS"])
        w.writerow(["date", "astreinte", "conges", "samedi_travaille", "repos_hebdo"])
        all_days = sorted(astreinte_days | conges_days | samedi_trav_days | repos_hebdo_days)
        for d in all_days:
            w.writerow([
                fmt(d),
                1 if d in astreinte_days else 0,
                1 if d in conges_days else 0,
                1 if d in samedi_trav_days else 0,
                1 if d in repos_hebdo_days else 0,
            ])

    return csv_out


def write_ics(path: Path, calname: str, events: list[tuple[str, str, date, date]]) -> None:
    lines: list[str] = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//excel_reader//FR",
        "CALSCALE:GREGORIAN",
        f"X-WR-CALNAME:{ics_escape(calname)}",
        f"X-WR-TIMEZONE:{ics_escape(ICS_TIMEZONE)}",
    ]

    dtstamp = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")

    for title, prefix, s, e in events:
        uid = ics_escape(block_uid(prefix, s, e))
        lines.extend([
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTAMP:{dtstamp}",
            f"DTSTART;VALUE=DATE:{s.strftime('%Y%m%d')}",
            f"DTEND;VALUE=DATE:{(e + timedelta(days=1)).strftime('%Y%m%d')}",
            f"SUMMARY:{ics_escape(title)}",
            "END:VEVENT",
        ])

    lines.append("END:VCALENDAR")
    path.write_text("\r\n".join(lines) + "\r\n", encoding="utf-8")


def export_ics(result: dict, timestamp: str, mode: str) -> list[Path]:
    """mode: 'per_category' ou 'all_in_one'"""
    out_paths: list[Path] = []

    ab = result["astreinte_blocks"]
    cb = result["conges_blocks"]
    sb = result["samedi_trav_blocks"]
    rb = result["repos_hebdo_blocks"]

    events_astreintes = [(ICS_EVENT_TITLE_ASTREINTE, "astreinte", s, e) for s, e in ab]
    events_conges = [(ICS_EVENT_TITLE_CONGES, "conges", s, e) for s, e in cb]
    events_samedi = [(ICS_EVENT_TITLE_SAMEDI_TRAV, "samedi_travaille", s, e) for s, e in sb]
    events_repos = [(ICS_EVENT_TITLE_REPOS_HEBDO, "repos_hebdo", s, e) for s, e in rb]

    if mode == "per_category":
        p1 = Path(f"{ICS_OUT_BASE}_astreintes_{timestamp}.ics")
        write_ics(p1, "Astreintes", events_astreintes)
        out_paths.append(p1)

        p2 = Path(f"{ICS_OUT_BASE}_conges_{timestamp}.ics")
        write_ics(p2, "Conges", events_conges)
        out_paths.append(p2)

        p3 = Path(f"{ICS_OUT_BASE}_samedi_travaille_{timestamp}.ics")
        write_ics(p3, "Samedi travaille", events_samedi)
        out_paths.append(p3)

        p4 = Path(f"{ICS_OUT_BASE}_repos_hebdo_{timestamp}.ics")
        write_ics(p4, "Repos hebdomadaire", events_repos)
        out_paths.append(p4)

    elif mode == "all_in_one":
        all_events = events_astreintes + events_conges + events_samedi + events_repos
        p = Path(f"{ICS_OUT_BASE}_tout_{timestamp}.ics")
        write_ics(p, ICS_CALNAME, all_events)
        out_paths.append(p)

    else:
        raise ValueError("mode doit etre 'per_category' ou 'all_in_one'")

    return out_paths


def run_pipeline(xlsx_path: Path, target_label: str, ics_mode: str, do_csv: bool, do_ics: bool) -> dict:
    ts = datetime.now().strftime("%d%m%y%H%M")
    result = compute_from_excel(xlsx_path=xlsx_path, target_label=target_label)

    csv_file = None
    ics_files: list[Path] = []

    if do_csv:
        csv_file = export_csv(result, ts)

    if do_ics:
        ics_files = export_ics(result, ts, ics_mode)

    # Console recap
    print("\n=== RECAP ===")
    print(f"Target: {target_label}")
    print(f"Astreintes : {len(result['astreinte_days'])} jours / {len(result['astreinte_blocks'])} blocs")
    print(f"Conges     : {len(result['conges_days'])} jours / {len(result['conges_blocks'])} blocs")
    print(f"Samedi trav: {len(result['samedi_trav_days'])} jours / {len(result['samedi_trav_blocks'])} blocs")
    print(f"Repos hebdo: {len(result['repos_hebdo_days'])} jours / {len(result['repos_hebdo_blocks'])} blocs")
    if result["unknown_cols"]:
        print(f"[WARN] Colonnes sans date exploitable: {len(result['unknown_cols'])}")

    if csv_file:
        print("CSV exporte :", csv_file.resolve())
    for p in ics_files:
        print("ICS exporte :", p.resolve())

    result["timestamp"] = ts
    result["csv_file"] = csv_file
    result["ics_files"] = ics_files
    return result


# =====================
# GUI (Tkinter)
# =====================

def launch_gui() -> None:
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog

    root = tk.Tk()
    root.title("Export planning (Excel -> CSV/ICS)")
    root.geometry("560x320")

    # Vars
    xlsx_var = tk.StringVar(value=str(DEFAULT_XLSX_PATH))
    target_var = tk.StringVar(value=DEFAULT_TARGET_LABEL)
    ics_mode_var = tk.StringVar(value=DEFAULT_ICS_MODE)
    do_csv_var = tk.BooleanVar(value=EXPORT_CSV)
    do_ics_var = tk.BooleanVar(value=EXPORT_ICS)

    def browse_xlsx():
        p = filedialog.askopenfilename(
            title="Choisir le fichier Excel",
            filetypes=[("Excel", "*.xlsx"), ("Tous fichiers", "*")],
            initialdir=str(BASE_DIR),
        )
        if p:
            xlsx_var.set(p)

    def run_clicked():
        try:
            xlsx_path = Path(xlsx_var.get()).expanduser().resolve()
            if not xlsx_path.exists():
                raise FileNotFoundError(f"Fichier introuvable: {xlsx_path}")

            target = target_var.get().strip()
            if not target:
                raise ValueError("TARGET_LABEL vide")

            mode = ics_mode_var.get()
            if mode not in ("per_category", "all_in_one"):
                raise ValueError("Mode calendrier invalide")

            run_pipeline(
                xlsx_path=xlsx_path,
                target_label=target,
                ics_mode=mode,
                do_csv=bool(do_csv_var.get()),
                do_ics=bool(do_ics_var.get()),
            )

            messagebox.showinfo("OK", "Export termine.\nRegarde la console pour les chemins.")

        except Exception as e:
            messagebox.showerror("Erreur", str(e))

    pad = {"padx": 10, "pady": 6}

    frm = ttk.Frame(root)
    frm.pack(fill="both", expand=True, **pad)

    # XLSX
    ttk.Label(frm, text="Fichier Excel (.xlsx)").grid(row=0, column=0, sticky="w")
    xlsx_entry = ttk.Entry(frm, textvariable=xlsx_var, width=55)
    xlsx_entry.grid(row=1, column=0, sticky="we")
    ttk.Button(frm, text="Parcourir...", command=browse_xlsx).grid(row=1, column=1, sticky="e", padx=8)

    # Target
    ttk.Label(frm, text="Target label (colonne A) ex: M09").grid(row=2, column=0, sticky="w")
    ttk.Entry(frm, textvariable=target_var, width=20).grid(row=3, column=0, sticky="w")

    # Options export
    opts = ttk.LabelFrame(frm, text="Exports")
    opts.grid(row=4, column=0, columnspan=2, sticky="we", pady=10)

    ttk.Checkbutton(opts, text="Exporter CSV", variable=do_csv_var).grid(row=0, column=0, sticky="w", padx=10, pady=6)
    ttk.Checkbutton(opts, text="Exporter calendrier (ICS)", variable=do_ics_var).grid(row=1, column=0, sticky="w", padx=10, pady=6)

    ttk.Label(opts, text="Mode calendrier:").grid(row=0, column=1, sticky="w", padx=10)
    ttk.Radiobutton(opts, text="1 ICS par categorie", value="per_category", variable=ics_mode_var).grid(row=1, column=1, sticky="w", padx=10)
    ttk.Radiobutton(opts, text="1 ICS tout-en-un", value="all_in_one", variable=ics_mode_var).grid(row=2, column=1, sticky="w", padx=10)

    # Run
    ttk.Button(frm, text="Lancer l'export", command=run_clicked).grid(row=5, column=0, sticky="w")

    # layout
    frm.columnconfigure(0, weight=1)

    root.mainloop()


if __name__ == "__main__":
    # Par defaut: GUI. Si tu preferes la console, commente launch_gui() et appelle run_pipeline(...)
    launch_gui()
