from openpyxl import load_workbook
from pathlib import Path
import csv

# =========================
# CONFIG
# =========================
BASE_DIR = Path(__file__).resolve().parent
XLSX_PATH = BASE_DIR / "ROW 2026 FR113 NextGen Schedule 4.1_.xlsx"
SHEET_NAME = None  # None = feuille active

START_ROW = 2      # première ligne technicien (à ajuster si besoin)
COL_ID = 1
COL_NAME = 2
START_DAY_COL = 3  # colonnes des jours = à partir de la 3 (après ID + nom)

# Ligne qui contient les jours de semaine ("Sat", ...)
ROW_WEEKDAY = 3

OUT_CSV_ASTREINTE = BASE_DIR / "classement_astreintes.csv"
OUT_CSV_SAMEDI = BASE_DIR / "classement_samedis_travailles.csv"

PREFIX_ID = "FR113"

TOKENS_STRONG = {"-14", "-H"}
TOKEN_H = "H"

SAT_TOKENS = {"Sat", "SAT", "Saturday"}  # adapte si besoin
SAT_WORK_TOKENS = {"-14", "8,5", "8.5"}


def norm(v) -> str:
    """Normalise la cellule en string propre."""
    if v is None:
        return ""
    s = str(v).strip()
    s = s.replace("–", "-").replace("—", "-")  # tirets exotiques

    # normaliser 8,5 / 8.5
    if s == "8.5":
        return "8.5"
    if s == "8,5":
        return "8,5"

    return s


def count_astreinte_days(day_values: list[str]) -> int:
    """
    Règle:
      - -14 ou -H -> compte
      - H -> compte si encadré par -14/-H à gauche et à droite
    """
    n = len(day_values)
    if n == 0:
        return 0

    is_strong = [v in TOKENS_STRONG for v in day_values]
    is_h = [v == TOKEN_H for v in day_values]

    prefix_strong = [False] * n
    seen = False
    for i in range(n):
        if is_strong[i]:
            seen = True
        prefix_strong[i] = seen

    suffix_strong = [False] * n
    seen = False
    for i in range(n - 1, -1, -1):
        if is_strong[i]:
            seen = True
        suffix_strong[i] = seen

    total = 0
    for i in range(n):
        if is_strong[i]:
            total += 1
        elif is_h[i]:
            has_left = prefix_strong[i - 1] if i - 1 >= 0 else False
            has_right = suffix_strong[i + 1] if i + 1 < n else False
            if has_left and has_right:
                total += 1
    return total


def main():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    # 1) Colonnes "samedi"
    saturday_cols = []
    for c in range(START_DAY_COL, max_col + 1):
        wd = norm(ws.cell(ROW_WEEKDAY, c).value)
        if wd in SAT_TOKENS:
            saturday_cols.append(c)

    # 2) Parcours techniciens
    astreinte_results = []
    saturday_results = []
    combined_results = []

    for r in range(START_ROW, max_row + 1):
        # IGNORER totalement la ligne 2
        if r == 2:
            continue
        # IGNORER toute ligne sans valeur en colonne A
        raw_id = ws.cell(r, COL_ID).value
        raw_name = ws.cell(r, COL_NAME).value

        # IGNORER toute ligne sans valeur en colonne A OU B
        if raw_id is None or str(raw_id).strip() == "" or raw_name is None or str(raw_name).strip() == "":
            continue

        # IGNORER aussi les lignes "header" (ex: "CE Name") ou celles où la colonne B est un lien
        raw_name_str = str(raw_name).strip()
        raw_name_low = raw_name_str.lower()
        is_header_name = raw_name_low == "ce name"
        is_link = (
            raw_name_low.startswith("http://")
            or raw_name_low.startswith("https://")
            or raw_name_low.startswith("www.")
            or "hyperlink(" in raw_name_low
            or raw_name_low.startswith("=hyperlink(")
        )
        if is_header_name or is_link:
            continue

        name = norm(raw_name)

        tech_id = f"{PREFIX_ID}{norm(raw_id)}"

        # astreintes
        day_values = [norm(ws.cell(r, c).value) for c in range(START_DAY_COL, max_col + 1)]
        astreinte_count = count_astreinte_days(day_values)

        # samedis travaillés
        sat_count = 0
        for c in saturday_cols:
            v = norm(ws.cell(r, c).value)
            if v in SAT_WORK_TOKENS:
                sat_count += 1

        astreinte_results.append((tech_id, name, astreinte_count))
        saturday_results.append((tech_id, name, sat_count))
        combined_results.append((tech_id, name, astreinte_count, sat_count))

    # 3) Tri
    astreinte_results.sort(key=lambda x: (-x[2], x[1].lower(), x[0]))
    saturday_results.sort(key=lambda x: (-x[2], x[1].lower(), x[0]))
    combined_results.sort(key=lambda x: (-x[2], -x[3], x[1].lower(), x[0]))

    # 4) Console
    nb_tech_total = len(astreinte_results)
    nb_tech_astreinte_non_nulle = sum(1 for _, _, c in astreinte_results if c > 0)

    print("=== Classement astreintes (jours) ===")
    print(f"Techniciens faisant de l'astreinte : {nb_tech_astreinte_non_nulle}/{nb_tech_total}")
    for tech_id, name, count in astreinte_results:
        print(f"{tech_id} | {name} | {count}")

    print("\n=== Classement samedis travaillés ===")
    for tech_id, name, count in saturday_results:
        print(f"{tech_id} | {name} | {count}")

    print("\n=== Classement combiné (astreintes, puis samedis) ===")
    for tech_id, name, a, s in combined_results:
        print(f"{tech_id} | {name} | astreintes={a} | samedis={s}")

    # 5) CSV
    def write_csv(path_str: str, header, rows):
        out_path = Path(path_str)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with out_path.open("w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(header)
            w.writerows(rows)
        print(f"CSV généré: {out_path}")

    write_csv(OUT_CSV_ASTREINTE, ["technicien_id", "nom", "jours_astreinte"], astreinte_results)
    write_csv(OUT_CSV_SAMEDI, ["technicien_id", "nom", "samedis_travailles"], saturday_results)

    # CSV combiné (pratique)
    out_combined = str(Path(OUT_CSV_ASTREINTE).with_name("classement_combine.csv"))
    write_csv(out_combined, ["technicien_id", "nom", "jours_astreinte", "samedis_travailles"], combined_results)


if __name__ == "__main__":
    main()
